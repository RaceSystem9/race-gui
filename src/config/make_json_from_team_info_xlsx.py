import json
from openpyxl import load_workbook

# 엑셀 파일 읽기
wb = load_workbook("team_info.xlsx", data_only=True)
ws = wb.active

teams = []


# 안전하게 로우의 값을 가져오는 헬퍼 함수
def get_val(row_tuple, index):
    return row_tuple[index] if index < len(row_tuple) else None


# 첫 번째 행은 제목(min_row=2부터 시작)
for row in ws.iter_rows(min_row=2, values_only=True):
    # 팀 번호가 없으면 스킵
    if not row or row[0] is None:
        continue

    # 팀원 목록 생성
    members = []

    for i in range(5):
        base = 5 + i * 4

        name = get_val(row, base)
        # 이름이 없으면 해당 팀원 정보는 스킵
        if name is None:
            continue

        members.append(
            {
                "name": name,
                "school": get_val(row, base + 1),
                "department": get_val(row, base + 2),
                "grade": get_val(row, base + 3),
            }
        )

    # 2번째 열(row[1]) 값이 '기권'이면 1, 아니면 0
    is_giveup = 1 if get_val(row, 1) == "기권" else 0
    
    # JSON 데이터 구조 생성
    team = {
        "team_no": int(row[0]),
        "giveup": is_giveup,  # "기권"이면 1, 아니면 0
        "team_name": get_val(row, 2),
        "school": get_val(row, 3),
        "num_of_members": len(members),
        "members": members,
    }

    teams.append(team)

# JSON 저장
with open("team_info.json", "w", encoding="utf-8") as f:
    json.dump(teams, f, ensure_ascii=False, indent=4)

print("team_info.xlsx 파일을 성공적으로 처리했습니다.")
print(f"총 {len(teams)}개 팀 정보를 team_info.json으로 저장했습니다.")